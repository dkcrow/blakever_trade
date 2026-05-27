#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
七星172交易记录持久化 - 生成xlsx格式的完整买卖记录表

输出: /workspace/blakever_trade/backtest/results_172/七星172_交易记录_2026.xlsx
"""

import os
import sys
import math
import json
import warnings
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

# 项目路径
sys.path.insert(0, str(Path(__file__).parent.parent))
from strategies.etf.seven_star_base import (
    LocalDataSource, Portfolio, ETF_POOL, ETF_NAMES, DEFENSIVE_ETF
)

START_DATE = '2026-01-01'
END_DATE = '2026-05-27'
INITIAL_CASH = 10000

# ================================================================
# 🎯 内联七星172引擎 (增强版: 记录动量得分)
# ================================================================

class SeveStar172WithScores:
    """七星172引擎, 额外记录每次交易的动量得分"""

    def __init__(self):
        self.lookback_days = 25
        self.holdings_num = 1
        self.min_money = 5000
        self.loss_threshold = 0.97
        self.min_score = 0
        self.max_score = 100.0
        self.protection_lookback = 1
        self.protection_threshold = 0.05
        self.short_lookback = 10
        self.short_threshold = 0.0
        self.volume_lookback = 5
        self.volume_threshold = 2
        self.volume_return_limit = 1
        self.profit_protection_sold_today = []

    def reset_daily(self):
        self.profit_protection_sold_today = []

    def check_profit_protection(self, code, current_price, hist_df, date):
        try:
            mask = hist_df.index < pd.Timestamp(date)
            hist_before = hist_df[mask]
            if len(hist_before) < self.protection_lookback:
                return False
            max_high = hist_before['high'].tail(self.protection_lookback).max()
            if max_high > 0 and current_price <= max_high * (1 - self.protection_threshold):
                return True
        except:
            pass
        return False

    def get_volume_ratio(self, hist_df, date):
        try:
            mask = hist_df.index <= pd.Timestamp(date)
            h = hist_df[mask]
            if len(h) < self.volume_lookback + 1:
                return None
            vols = h['volume'].tail(self.volume_lookback + 1)
            avg = vols.iloc[:-1].mean()
            cur = vols.iloc[-1]
            if avg > 0:
                r = cur / avg
                if r > self.volume_threshold:
                    return r
        except:
            pass
        return None

    def get_ranked_etfs(self, all_etf_data, current_prices, date):
        """计算排名, 返回带得分的列表"""
        etf_metrics = []
        for code in ETF_POOL:
            df = all_etf_data.get(code)
            if df is None:
                continue
            mask = df.index <= pd.Timestamp(date)
            hist = df[mask]
            if len(hist) < self.lookback_days:
                continue

            close_arr = hist['close'].values
            cur_price = current_prices.get(code, 0)
            if cur_price <= 0:
                continue

            close_full = np.append(close_arr, cur_price)

            # 盈利保护
            if self.check_profit_protection(code, cur_price, df, date):
                continue

            # 成交量过滤
            if self.get_volume_ratio(df, date) is not None:
                ann = self._calc_annual(close_full, self.lookback_days)
                if ann > self.volume_return_limit:
                    continue

            # 短期动量
            if len(close_full) >= self.short_lookback + 1:
                sr = close_full[-1] / close_full[-(self.short_lookback + 1)] - 1
                sa = (1 + sr) ** (250 / self.short_lookback) - 1
            else:
                sa = 0

            if sa < self.short_threshold:
                continue

            # 长期动量 + R²
            recent = close_full[-(self.lookback_days + 1):]
            y = np.log(np.maximum(recent, 1e-10))
            x = np.arange(len(y))
            w = np.linspace(1, 2, len(y))
            slope, intercept = np.polyfit(x, y, 1, w=w)
            ann_ret = math.exp(slope * 250) - 1

            ss_res = np.sum(w * (y - (slope * x + intercept)) ** 2)
            ss_tot = np.sum(w * (y - np.mean(y)) ** 2)
            r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0

            score = ann_ret * r2

            # 近3日跌幅
            if len(close_full) >= 4:
                d1 = close_full[-1] / close_full[-2]
                d2 = close_full[-2] / close_full[-3]
                d3 = close_full[-3] / close_full[-4]
                if min(d1, d2, d3) < self.loss_threshold:
                    continue

            if not (self.min_score < score < self.max_score):
                continue

            etf_metrics.append({
                'code': code,
                'name': ETF_NAMES.get(code, code),
                'annualized': ann_ret,
                'r2': r2,
                'score': score,
                'price': cur_price,
            })

        etf_metrics.sort(key=lambda x: x['score'], reverse=True)
        return etf_metrics

    def _calc_annual(self, prices, lb):
        recent = prices[-(lb + 1):]
        y = np.log(np.maximum(recent, 1e-10))
        x = np.arange(len(y))
        w = np.linspace(1, 2, len(y))
        slope, _ = np.polyfit(x, y, 1, w=w)
        return math.exp(slope * 250) - 1


# ================================================================
# 🔄 回测执行
# ================================================================

def run_backtest():
    ds = LocalDataSource()
    engine = SeveStar172WithScores()

    print("=" * 60)
    print("七星172交易记录生成")
    print(f"区间: {START_DATE} ~ {END_DATE}, 初始: {INITIAL_CASH:,}")
    print("=" * 60)

    all_data = ds.load_all_etfs(START_DATE, END_DATE)
    print(f"加载 {len(all_data)} 只ETF")

    trade_dates = ds.get_trade_dates(START_DATE, END_DATE)
    print(f"交易日: {len(trade_dates)} 天")

    portfolio = Portfolio(initial_cash=INITIAL_CASH, commission_rate=0.0002, min_commission=5)

    # 增强的交易记录: 每条附带动量得分
    enhanced_records = []

    for i, td in enumerate(trade_dates):
        td_ts = pd.Timestamp(td)

        current_prices = {}
        for code, df in all_data.items():
            mask = df.index <= td_ts
            if mask.any():
                current_prices[code] = float(df.loc[mask, 'close'].iloc[-1])

        portfolio.update_prices(current_prices)

        # 清空黑名单
        engine.reset_daily()

        # 盈利保护检查
        for code in list(portfolio.get_position_codes()):
            if code not in all_data:
                continue
            cp = current_prices.get(code, 0)
            if cp <= 0:
                continue
            if engine.check_profit_protection(code, cp, all_data[code], td):
                if portfolio.sell_all(code, cp, td, reason='盈利保护'):
                    engine.profit_protection_sold_today.append(code)
                    # 记录卖出 (得分用最后缓存或重新计算)
                    ranked = engine.get_ranked_etfs(all_data, current_prices, td)
                    score = next((m['score'] for m in ranked if m['code'] == code), None)
                    enhanced_records.append({
                        '交易日期': str(td),
                        'ETF名称': ETF_NAMES.get(code, code),
                        'ETF代码': code,
                        '方向': '卖出',
                        '成交价格': round(cp, 4),
                        '综合动量得分': round(score, 4) if score is not None else '',
                        '交易理由': '盈利保护(回撤超5%)',
                    })

        # 卖出操作
        ranked = engine.get_ranked_etfs(all_data, current_prices, td)
        target_codes = [m['code'] for m in ranked[:engine.holdings_num] if m['score'] >= engine.min_score]

        for sec in list(portfolio.get_position_codes()):
            if sec not in target_codes:
                cp = current_prices.get(sec, 0)
                if cp <= 0:
                    continue
                if portfolio.sell_all(sec, cp, td, reason='调出目标'):
                    # 获取卖出时的得分 (可能在排名中或已被过滤)
                    sell_score = next((m['score'] for m in ranked if m['code'] == sec), None)
                    if sell_score is None:
                        # 不在当前排名中, 可能是参数已变, 用N/A
                        sell_score = 'N/A'
                    enhanced_records.append({
                        '交易日期': str(td),
                        'ETF名称': ETF_NAMES.get(sec, sec),
                        'ETF代码': sec,
                        '方向': '卖出',
                        '成交价格': round(cp, 4),
                        '综合动量得分': round(sell_score, 4) if isinstance(sell_score, (int, float)) else sell_score,
                        '交易理由': '调出目标(排名下降)',
                    })

        # 买入操作
        target_etfs = []
        for m in ranked:
            if len(target_etfs) >= engine.holdings_num:
                break
            etf = m['code']

            # 二次检查
            if engine.check_profit_protection(etf, m['price'], all_data.get(etf, pd.DataFrame()), td):
                continue

            # 黑名单
            if etf in engine.profit_protection_sold_today:
                continue

            target_etfs.append(etf)

        if not target_etfs:
            target_etfs = [DEFENSIVE_ETF]

        total_val = portfolio.total_value
        target_per = total_val / len(target_etfs)

        for idx, etf in enumerate(target_etfs):
            cp = current_prices.get(etf, 0)
            if cp <= 0:
                continue

            current_val = 0
            if etf in portfolio.positions:
                pos = portfolio.positions[etf]
                if pos['shares'] > 0:
                    current_val = pos['shares'] * pos.get('last_price', pos['cost_price'])

            diff = target_per - current_val
            if abs(diff) < target_per * 0.05 and current_val > 0:
                continue

            if diff > 0:
                target_amt = int(diff / cp // 100) * 100
                if target_amt <= 0 and diff > engine.min_money:
                    target_amt = 100
                if target_amt * cp >= engine.min_money:
                    if portfolio.buy(etf, target_amt, cp, td, reason=f'排名{idx+1}/{len(ranked)}'):
                        # 获取买入得分
                        buy_score = next((m['score'] for m in ranked if m['code'] == etf), None)
                        reason_str = f'动量排名第{idx+1}/{len(ranked)}'
                        if etf == DEFENSIVE_ETF:
                            reason_str = '防御模式(无合格标的)'
                            buy_score = 'N/A'
                        enhanced_records.append({
                            '交易日期': str(td),
                            'ETF名称': ETF_NAMES.get(etf, etf),
                            'ETF代码': etf,
                            '方向': '买入',
                            '成交价格': round(cp, 4),
                            '综合动量得分': round(buy_score, 4) if isinstance(buy_score, (int, float)) else buy_score,
                            '交易理由': reason_str,
                        })

        portfolio.record_daily_value(td)

    print(f"\n回测完成!")
    print(f"总交易: {len(enhanced_records)} 条")
    print(f"买入: {sum(1 for r in enhanced_records if r['方向']=='买入')} 条")
    print(f"卖出: {sum(1 for r in enhanced_records if r['方向']=='卖出')} 条")
    print(f"初始: {INITIAL_CASH:,.2f}, 最终: {portfolio.total_value:,.2f}")
    print(f"收益率: {portfolio.total_returns*100:+.2f}%")

    return enhanced_records, portfolio


# ================================================================
# 📊 生成xlsx
# ================================================================

def create_xlsx(records, portfolio):
    output_dir = Path(__file__).parent.parent / 'backtest' / 'results_172'
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / '七星172_交易记录_2026.xlsx'

    df = pd.DataFrame(records)
    # 确保列顺序
    columns = ['交易日期', 'ETF名称', 'ETF代码', '方向', '成交价格', '综合动量得分', '交易理由']
    df = df[columns]

    # 写入Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '交易记录'

    # 颜色和样式
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    cell_font = Font(name='微软雅黑', size=10)
    buy_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # 浅绿
    sell_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid') # 浅橙
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 写标题行
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 写数据行
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        direction = row['方向']
        fill = buy_fill if direction == '买入' else sell_fill

        for col_idx, col_name in enumerate(columns, 1):
            val = row[col_name]
            # 格式化
            if col_name == '成交价格' and isinstance(val, (int, float)):
                val = round(val, 4)
            elif col_name == '综合动量得分':
                if isinstance(val, (int, float)):
                    val = round(val, 4)
                elif val == '' or val == 'N/A':
                    val = 'N/A'

            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = cell_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if col_idx != 7 else 'left', vertical='center')
            cell.border = thin_border

    # 列宽
    col_widths = [14, 24, 14, 8, 12, 16, 40]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # 冻结首行
    ws.freeze_panes = 'A2'

    # ===== 汇总Sheet =====
    ws2 = wb.create_sheet('回测摘要')
    summary_data = [
        ['指标', '数值'],
        ['策略', '七星172 (GLM5修复版)'],
        ['回测区间', f'{START_DATE} ~ {END_DATE}'],
        ['初始资金', f'{INITIAL_CASH:,.0f} 元'],
        ['最终资产', f'{portfolio.total_value:,.2f} 元'],
        ['总收益率', f'{portfolio.total_returns*100:+.2f}%'],
        ['总交易次数', f'{len(records)} 条'],
        ['买入笔数', f"{sum(1 for r in records if r['方向']=='买入')} 条"],
        ['卖出笔数', f"{sum(1 for r in records if r['方向']=='卖出')} 条"],
        ['佣金费率', '0.02% (双边)'],
        ['防御ETF', f'{DEFENSIVE_ETF} {ETF_NAMES.get(DEFENSIVE_ETF)}'],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = cell_font
            cell.alignment = Alignment(horizontal='center' if col_idx == 1 else 'left')
            cell.border = thin_border

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 40
    ws2.freeze_panes = 'A2'

    # 保存
    wb.save(output_path)
    print(f"\n📄 交易记录已保存: {output_path}")
    return output_path, df


# ================================================================
# 🏃 主入口
# ================================================================

if __name__ == '__main__':
    records, portfolio = run_backtest()
    path, df = create_xlsx(records, portfolio)

    # 展示最近20条
    print("\n" + "=" * 100)
    print("最近 20 条交易记录")
    print("=" * 100)
    print(f"{'日期':<12} {'ETF名称':<20} {'代码':<12} {'方向':<6} {'价格':>8} {'动量得分':>10} {'交易理由':<30}")
    print("-" * 100)

    for r in records[-20:]:
        score_str = f"{r['综合动量得分']:>10.4f}" if isinstance(r['综合动量得分'], (int, float)) else f"{r['综合动量得分']:>10}"
        print(f"{r['交易日期']:<12} {r['ETF名称']:<20} {r['ETF代码']:<12} {r['方向']:<6} {r['成交价格']:>8.4f} {score_str} {r['交易理由']:<30}")

    print("-" * 100)
    print(f"\n📄 完整文件: {path}")
